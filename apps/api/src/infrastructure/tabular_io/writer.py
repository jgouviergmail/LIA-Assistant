"""Render a :class:`WorkbookSpec` and its rows into an ``.xlsx`` file.

The layout is designed so the file survives being edited by hand and read back
without ambiguity:

- **row 1** carries the technical column keys and is hidden. The reader matches
  on it, so reordering, hiding or inserting columns in Excel is harmless, and
  changing the interface language between export and import changes nothing.
- **row 2** carries translated labels, coloured by block so 25+ columns stay
  navigable without collapsible groups (which Excel disables on a protected
  sheet).
- **row 3 onward** is data. The key column and both header rows are frozen.

Two openpyxl attributes used here are inverted, and both are load-bearing;
``test_writer.py`` pins them against the emitted XML rather than the Python API.
"""

from __future__ import annotations

import io
from collections.abc import Mapping, Sequence
from decimal import Decimal
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from src.domains.document_generation.sanitize import neutralize_formula
from src.infrastructure.tabular_io.spec import ColumnSpec, SheetSpec, WorkbookSpec

#: First row holding data. Rows 1 and 2 are the technical keys and the labels.
DATA_TOP_ROW = 3

#: Blank rows pre-formatted below the data so dropdowns and number formats also
#: apply to rows the administrator adds. ``max_row`` is inflated as a result —
#: the reader never derives a count from it.
SPARE_ROWS = 200

#: Referential holding the localized yes/no words, always present.
BOOLEAN_REFERENTIAL = "BOOL"

_BLOCK_FILLS: Mapping[str, str] = {
    "identity": "DDEBF7",
    "state": "FFF2CC",
    "capabilities": "E2EFDA",
    "sampling": "FCE4D6",
    "reasoning": "EDEDED",
    "pricing": "D9E1F2",
    "slots": "FFF2CC",
    "diagnostics": "F2F2F2",
    "default": "F2F2F2",
}

_METADATA_VERSION_KEY = "sheet_schema_version"


def build_workbook(
    spec: WorkbookSpec,
    data: Mapping[str, Sequence[Mapping[str, Any]]],
    *,
    notice: Sequence[str],
    labels: Mapping[str, str],
    metadata: Mapping[str, str],
) -> bytes:
    """Render the workbook described by ``spec``.

    Args:
        spec: Declarative description of the sheets and their referentials.
        data: Rows per sheet name. A sheet absent from the mapping is written
            with its headers only.
        notice: Already translated lines of the usage notice.
        labels: Translated strings by i18n key. A missing key falls back to the
            column key, so a translation gap never yields a blank header.
        metadata: Extra key/value pairs recorded on the metadata sheet
            alongside the schema version.

    Returns:
        The ``.xlsx`` file as bytes.
    """
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    _write_notice(workbook, notice, labels)
    for sheet_spec in spec.sheets:
        _write_data_sheet(workbook, sheet_spec, data.get(sheet_spec.name, ()), labels)
    _write_referentials(workbook, spec, labels)
    for sheet_spec in spec.sheets:
        _attach_validations(
            workbook[sheet_spec.name], sheet_spec, len(data.get(sheet_spec.name, ()))
        )
    _write_metadata(workbook, spec, metadata, labels)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _label(labels: Mapping[str, str], key: str, fallback: str) -> str:
    """Resolve a translated label, never returning an empty string."""
    return labels.get(key) or fallback


def _write_notice(
    workbook: openpyxl.Workbook, notice: Sequence[str], labels: Mapping[str, str]
) -> None:
    sheet = workbook.create_sheet(_label(labels, "sheet.notice", "Notice"))
    for index, line in enumerate(notice, start=1):
        sheet.cell(row=index, column=1, value=line)
    if notice:
        sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
    sheet.column_dimensions["A"].width = 100
    sheet.protection.sheet = True


def _cell_value(column: ColumnSpec, raw: Any, labels: Mapping[str, str]) -> Any:
    """Convert a domain value into what the cell must hold.

    ``None`` and the empty string both become an empty cell: Excel reads an
    empty cell back as ``None``, so writing ``""`` would make every such row
    report a phantom change on the next import.
    """
    if raw is None or raw == "":
        return None
    if column.kind == "boolean":
        return _label(
            labels, "boolean.true" if raw else "boolean.false", "TRUE" if raw else "FALSE"
        )
    if column.kind == "decimal":
        return float(raw if isinstance(raw, Decimal) else Decimal(str(raw)))
    if column.kind == "integer":
        return int(raw)
    if isinstance(raw, str):
        # A stored value starting with = + - @ would otherwise be evaluated by
        # Excel when the file is opened (OWASP spreadsheet injection).
        return neutralize_formula(raw)
    return raw


def _write_data_sheet(
    workbook: openpyxl.Workbook,
    spec: SheetSpec,
    rows: Sequence[Mapping[str, Any]],
    labels: Mapping[str, str],
) -> None:
    sheet = workbook.create_sheet(spec.name)

    for index, column in enumerate(spec.columns, start=1):
        letter = get_column_letter(index)
        key_cell = sheet.cell(row=1, column=index, value=column.key)
        key_cell.font = Font(size=8, color="808080")
        label_cell = sheet.cell(
            row=2, column=index, value=_label(labels, column.label_key, column.key)
        )
        label_cell.font = Font(bold=True, size=9, color="000000" if column.editable else "7F7F7F")
        label_cell.fill = PatternFill("solid", fgColor=_BLOCK_FILLS.get(column.block, "F2F2F2"))
        label_cell.alignment = Alignment(wrap_text=True, vertical="center")
        sheet.column_dimensions[letter].width = column.width
        sheet.column_dimensions[letter].hidden = column.hidden

    last_row = DATA_TOP_ROW + max(len(rows), 1) + SPARE_ROWS
    for offset, row in enumerate(rows):
        for index, column in enumerate(spec.columns, start=1):
            sheet.cell(
                row=DATA_TOP_ROW + offset,
                column=index,
                value=_cell_value(column, row.get(column.key), labels),
            )

    # Formatting and unlocking cover the spare rows too, so a line the
    # administrator adds behaves exactly like an exported one.
    for index, column in enumerate(spec.columns, start=1):
        for row_number in range(DATA_TOP_ROW, last_row + 1):
            cell = sheet.cell(row=row_number, column=index)
            if column.kind == "decimal" and column.decimals is not None:
                cell.number_format = "0." + "0" * column.decimals
            cell.protection = Protection(locked=not column.editable)

    sheet.freeze_panes = f"{get_column_letter(2)}{DATA_TOP_ROW}"
    sheet.auto_filter.ref = f"A2:{get_column_letter(len(spec.columns))}{last_row}"
    sheet.row_dimensions[1].hidden = True
    _protect(sheet)


def _protect(sheet: Worksheet) -> None:
    """Lock the structure while leaving the administrator's tools working.

    Every attribute below means "this action is BLOCKED" when true, which is
    the opposite of how it reads. Left at their defaults they would forbid
    inserting a row — that is, forbid adding a model. No password: it is
    bypassed in seconds and would only create false confidence.
    """
    sheet.protection.sheet = True
    for allowed in (
        "autoFilter",
        "sort",
        "insertRows",
        "deleteRows",
        "formatCells",
        "formatColumns",
        "formatRows",
    ):
        setattr(sheet.protection, allowed, False)
    sheet.protection.selectLockedCells = False
    sheet.protection.selectUnlockedCells = False


def _write_referentials(
    workbook: openpyxl.Workbook, spec: WorkbookSpec, labels: Mapping[str, str]
) -> None:
    sheet = workbook.create_sheet(_label(labels, "sheet.referentials", "Referentials"))
    lists: dict[str, Sequence[str]] = dict(spec.referentials)
    lists[BOOLEAN_REFERENTIAL] = (
        _label(labels, "boolean.true", "TRUE"),
        _label(labels, "boolean.false", "FALSE"),
    )

    for index, (name, values) in enumerate(lists.items(), start=1):
        letter = get_column_letter(index)
        sheet.cell(row=1, column=index, value=name).font = Font(bold=True)
        for offset, value in enumerate(values, start=2):
            sheet.cell(row=offset, column=index, value=value)
        workbook.defined_names.add(
            DefinedName(
                f"LST_{name}",
                attr_text=f"'{sheet.title}'!${letter}$2:${letter}${1 + len(values)}",
            )
        )
    sheet.sheet_state = "hidden"
    sheet.protection.sheet = True


def _attach_validations(sheet: Worksheet, spec: SheetSpec, row_count: int) -> None:
    """Attach dropdowns and numeric bounds to the editable columns."""
    last_row = DATA_TOP_ROW + max(row_count, 1) + SPARE_ROWS

    for index, column in enumerate(spec.columns, start=1):
        if not column.editable:
            continue
        letter = get_column_letter(index)
        cells = f"{letter}{DATA_TOP_ROW}:{letter}{last_row}"

        referential = column.referential or (
            BOOLEAN_REFERENTIAL if column.kind == "boolean" else None
        )
        if referential:
            # showDropDown is inverted: leaving it False is what SHOWS the arrow.
            listing = DataValidation(
                type="list",
                formula1=f"=LST_{referential}",
                allow_blank=True,
                showDropDown=False,
                showErrorMessage=True,
            )
            sheet.add_data_validation(listing)
            listing.add(cells)

        if column.kind in {"decimal", "integer"} and column.minimum is not None:
            bound = DataValidation(
                type="decimal" if column.kind == "decimal" else "whole",
                operator="greaterThanOrEqual",
                formula1=str(column.minimum),
                showErrorMessage=True,
            )
            sheet.add_data_validation(bound)
            bound.add(cells)


def _write_metadata(
    workbook: openpyxl.Workbook,
    spec: WorkbookSpec,
    metadata: Mapping[str, str],
    labels: Mapping[str, str],
) -> None:
    sheet = workbook.create_sheet(_label(labels, "sheet.metadata", "Metadata"))
    pairs: list[tuple[str, str]] = [(_METADATA_VERSION_KEY, str(spec.schema_version))]
    pairs += [(key, value) for key, value in metadata.items()]
    for index, (key, value) in enumerate(pairs, start=1):
        sheet.cell(row=index, column=1, value=key).font = Font(bold=True)
        sheet.cell(row=index, column=2, value=value)
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 46
    sheet.protection.sheet = True
