"""Parse a hand-edited workbook back into typed rows.

Every rule below closes a trap measured against real Excel on 2026-08-18, not
an imagined one:

- the workbook is opened with ``data_only=False``. ``data_only=True`` returns
  ``None`` for a formula cell on a file Excel never opened, so the outcome
  would depend on which tool last saved it. A cell holding a formula is refused
  outright — deterministic, and it tells the administrator exactly what to fix.
- columns are resolved by the **technical key** in the hidden first row, so
  reordering, hiding or adding columns changes nothing.
- fully-empty rows are skipped and no count is ever derived from ``max_row``:
  the writer pre-formats spare rows, which inflates it to hundreds.
- numbers are coerced from what Excel really hands back — ``'0,7'``,
  ``'1 234,50'`` with a non-breaking space, padded text — and refused rather
  than rounded when they carry more decimals than the column declares.
- an empty string and an absent cell are the same value; treating them apart
  made 122 rows of a real catalogue report a change that did not exist.
"""

from __future__ import annotations

import io
import re
import zipfile
from collections.abc import Mapping, Sequence
from decimal import Decimal, InvalidOperation
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.infrastructure.archives.zip_budget import ZipBudgetExceeded, enforce_zip_budgets
from src.infrastructure.tabular_io.report import (
    CellIssue,
    IssueCode,
    ParsedRow,
    ParsedWorkbook,
)
from src.infrastructure.tabular_io.spec import ColumnSpec, SheetSpec, WorkbookSpec
from src.infrastructure.tabular_io.writer import DATA_TOP_ROW

_TRUE_WORDS = frozenset({"VRAI", "TRUE", "OUI", "YES", "1", "X", "JA", "SI", "是"})
_FALSE_WORDS = frozenset({"FAUX", "FALSE", "NON", "NO", "0", "NEIN", "否"})

_NON_BREAKING_SPACES = (" ", " ", " ")
_HHMM = re.compile(r"^(?:[01]\d|2[0-3]):[0-5]\d$")
_METADATA_VERSION_KEY = "sheet_schema_version"


def parse_workbook(
    spec: WorkbookSpec,
    content: bytes,
    *,
    max_rows: int,
    max_files: int,
    max_decompressed_bytes: int,
) -> ParsedWorkbook:
    """Read ``content`` as the workbook described by ``spec``.

    Args:
        spec: The declaration the file is expected to follow.
        content: Raw uploaded bytes.
        max_rows: Largest number of data rows accepted per sheet.
        max_files: Largest number of archive members accepted.
        max_decompressed_bytes: Largest expanded archive size accepted.

    Returns:
        Typed rows per sheet and every issue found. Nothing raises: a caller
        renders the issues, it does not catch exceptions.
    """
    issues: list[CellIssue] = []

    workbook = _open(content, max_files, max_decompressed_bytes, issues)
    if workbook is None:
        return ParsedWorkbook(sheets={}, issues=tuple(issues))

    try:
        _check_schema_version(workbook, spec, issues)

        sheets: dict[str, Sequence[ParsedRow]] = {}
        for sheet_spec in spec.sheets:
            if sheet_spec.name not in workbook.sheetnames:
                issues.append(CellIssue(IssueCode.SHEET_MISSING, sheet=sheet_spec.name))
                continue
            sheets[sheet_spec.name] = _parse_sheet(
                workbook[sheet_spec.name], sheet_spec, spec, max_rows, issues
            )
        return ParsedWorkbook(sheets=sheets, issues=tuple(issues))
    finally:
        workbook.close()


def _open(
    content: bytes,
    max_files: int,
    max_decompressed_bytes: int,
    issues: list[CellIssue],
) -> openpyxl.Workbook | None:
    """Validate the archive budgets, then load the workbook."""
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            members = [info for info in archive.infolist() if not info.is_dir()]
            enforce_zip_budgets(
                members,
                max_files=max_files,
                max_decompressed_bytes=max_decompressed_bytes,
            )
    except zipfile.BadZipFile:
        issues.append(CellIssue(IssueCode.NOT_A_WORKBOOK))
        return None
    except ZipBudgetExceeded as exc:
        issues.append(
            CellIssue(
                IssueCode.ARCHIVE_TOO_LARGE,
                params={
                    "reason": exc.reason,
                    "limit": str(exc.limit),
                    "measured": str(exc.measured),
                },
            )
        )
        return None

    try:
        # data_only=False on purpose: see the module docstring.
        return openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    except (zipfile.BadZipFile, KeyError, ValueError):
        issues.append(CellIssue(IssueCode.NOT_A_WORKBOOK))
        return None


def _check_schema_version(
    workbook: openpyxl.Workbook, spec: WorkbookSpec, issues: list[CellIssue]
) -> None:
    """Refuse a workbook built for another schema.

    The metadata sheet is located by its content rather than its name, which is
    localized. A workbook without one is tolerated — only a *wrong* version is
    refused, so a file rebuilt by hand keeps working.
    """
    for sheet in workbook.worksheets:
        for row in range(1, min(sheet.max_row, 20) + 1):
            if sheet.cell(row=row, column=1).value != _METADATA_VERSION_KEY:
                continue
            found = str(sheet.cell(row=row, column=2).value or "").strip()
            if found and found != str(spec.schema_version):
                issues.append(
                    CellIssue(
                        IssueCode.SCHEMA_VERSION_MISMATCH,
                        sheet=sheet.title,
                        params={"expected": str(spec.schema_version), "found": found},
                    )
                )
            return


def _column_index(sheet: Worksheet, spec: SheetSpec, issues: list[CellIssue]) -> dict[str, int]:
    """Map each declared column key to its 1-based worksheet column.

    Resolution is by key, never by position. A column the spec does not know is
    ignored — an administrator may keep working notes beside the data.
    """
    found: dict[str, int] = {}
    for index in range(1, sheet.max_column + 1):
        raw = sheet.cell(row=1, column=index).value
        if isinstance(raw, str) and raw.strip():
            found.setdefault(raw.strip(), index)

    index_by_key: dict[str, int] = {}
    for column in spec.columns:
        if column.key in found:
            index_by_key[column.key] = found[column.key]
        elif column.editable:
            issues.append(CellIssue(IssueCode.COLUMN_MISSING, sheet=spec.name, column=column.key))
    return index_by_key


def _parse_sheet(
    sheet: Worksheet,
    spec: SheetSpec,
    workbook_spec: WorkbookSpec,
    max_rows: int,
    issues: list[CellIssue],
) -> tuple[ParsedRow, ...]:
    index_by_key = _column_index(sheet, spec, issues)
    if spec.key_column not in index_by_key:
        return ()

    rows: list[ParsedRow] = []
    seen_keys: dict[str, int] = {}

    for row_number in range(DATA_TOP_ROW, sheet.max_row + 1):
        raw = {
            key: sheet.cell(row=row_number, column=index).value
            for key, index in index_by_key.items()
        }
        if _is_blank(raw.values()):
            continue

        if len(rows) >= max_rows:
            issues.append(
                CellIssue(
                    IssueCode.TOO_MANY_ROWS,
                    sheet=spec.name,
                    params={"limit": str(max_rows)},
                )
            )
            break

        values, derived = _parse_row(raw, spec, workbook_spec, index_by_key, row_number, issues)
        key = values.get(spec.key_column)
        key_text = str(key) if key not in (None, "") else None

        if key_text is None:
            issues.append(
                CellIssue(
                    IssueCode.KEY_MISSING,
                    sheet=spec.name,
                    cell=_coordinate(index_by_key[spec.key_column], row_number),
                    column=spec.key_column,
                )
            )
        elif spec.key_is_unique and key_text in seen_keys:
            issues.append(
                CellIssue(
                    IssueCode.DUPLICATE_KEY,
                    sheet=spec.name,
                    cell=_coordinate(index_by_key[spec.key_column], row_number),
                    column=spec.key_column,
                    params={"key": key_text, "first_row": str(seen_keys[key_text])},
                )
            )
        else:
            seen_keys[key_text] = row_number

        rows.append(ParsedRow(row_number=row_number, key=key_text, values=values, derived=derived))

    return tuple(rows)


def _parse_row(
    raw: Mapping[str, Any],
    spec: SheetSpec,
    workbook_spec: WorkbookSpec,
    index_by_key: Mapping[str, int],
    row_number: int,
    issues: list[CellIssue],
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Split a row into what the admin typed and what the export computed.

    A read-only cell never raises an issue: the administrator cannot fix a
    value they cannot edit, so complaining about it would be noise. Its
    coercion failures simply yield ``None``.
    """
    values: dict[str, Any] = {}
    derived: dict[str, Any] = {}
    for column in spec.columns:
        if column.key not in index_by_key:
            continue
        cell = _coordinate(index_by_key[column.key], row_number)
        if column.editable:
            values[column.key] = _coerce(
                raw.get(column.key), column, workbook_spec, spec.name, cell, issues
            )
        else:
            derived[column.key] = _coerce(
                raw.get(column.key), column, workbook_spec, spec.name, cell, []
            )
    return values, derived


def _is_blank(values: Any) -> bool:
    """A row is empty when every declared cell is empty or blank text."""
    return all(value is None or (isinstance(value, str) and not value.strip()) for value in values)


def _coordinate(column_index: int, row_number: int) -> str:
    return f"{get_column_letter(column_index)}{row_number}"


def _coerce(
    raw: Any,
    column: ColumnSpec,
    workbook_spec: WorkbookSpec,
    sheet_name: str,
    cell: str,
    issues: list[CellIssue],
) -> Any:
    """Turn one cell into its declared type, reporting rather than guessing."""

    def report(code: IssueCode, **params: str) -> None:
        issues.append(
            CellIssue(code, sheet=sheet_name, cell=cell, column=column.key, params=params)
        )

    if isinstance(raw, str) and raw.startswith("="):
        report(IssueCode.FORMULA_REJECTED, value=raw[:80])
        return None

    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return None

    if column.kind == "boolean":
        return _coerce_boolean(raw, report)
    if column.kind in {"decimal", "integer"}:
        return _coerce_number(raw, column, report)
    if column.kind == "time_hhmm":
        text = str(raw).strip()
        if not _HHMM.match(text):
            report(IssueCode.NOT_A_TIME, value=text[:20])
            return None
        return text
    if column.kind == "enum":
        return _coerce_enum(str(raw), column, workbook_spec, report)
    if column.kind == "enum_list":
        return _coerce_enum_list(str(raw), column, workbook_spec, report)
    return str(raw).strip()


def _coerce_boolean(raw: Any, report: Any) -> bool | None:
    if isinstance(raw, bool):
        return raw
    text = str(raw).strip().upper()
    if text in _TRUE_WORDS:
        return True
    if text in _FALSE_WORDS:
        return False
    report(IssueCode.NOT_A_BOOLEAN, value=text[:20])
    return None


def _coerce_number(raw: Any, column: ColumnSpec, report: Any) -> Decimal | int | None:
    if isinstance(raw, bool):
        report(IssueCode.NOT_A_NUMBER, value=str(raw))
        return None

    if isinstance(raw, str):
        text = str(raw).strip()
        for space in _NON_BREAKING_SPACES:
            text = text.replace(space, "")
        text = text.replace(",", ".")
        try:
            value = Decimal(text)
        except InvalidOperation:
            report(IssueCode.NOT_A_NUMBER, value=text[:20])
            return None
    else:
        value = Decimal(str(raw))

    # Decimal happily parses "NaN" and "inf". Both would slip past the minimum
    # check below (every comparison with NaN is false) and then break the scale
    # computation, turning a typo into a 500 instead of a cell-level issue.
    if not value.is_finite():
        report(IssueCode.NOT_A_NUMBER, value=str(value))
        return None

    if column.minimum is not None and value < column.minimum:
        report(IssueCode.OUT_OF_RANGE, value=str(value), minimum=str(column.minimum))
        return None

    if column.kind == "integer":
        if value != value.to_integral_value():
            report(IssueCode.NOT_A_NUMBER, value=str(value))
            return None
        return int(value)

    # ``exponent`` is a str sentinel for non-finite values; those were already
    # refused above, and narrowing on int keeps that guarantee visible to the
    # type checker rather than resting on the earlier branch.
    exponent = value.as_tuple().exponent
    if not isinstance(exponent, int):
        report(IssueCode.NOT_A_NUMBER, value=str(value))
        return None
    if column.decimals is not None and -exponent > column.decimals:
        report(IssueCode.TOO_MANY_DECIMALS, value=str(value), decimals=str(column.decimals))
        return None
    return value


def _referential(column: ColumnSpec, workbook_spec: WorkbookSpec) -> Sequence[str]:
    return workbook_spec.referentials[column.referential or ""]


def _coerce_enum(
    raw: str, column: ColumnSpec, workbook_spec: WorkbookSpec, report: Any
) -> str | None:
    """Match a referential value, tolerating case and surrounding spaces."""
    text = raw.strip()
    for allowed in _referential(column, workbook_spec):
        if text.casefold() == allowed.casefold():
            return allowed
    report(IssueCode.VALUE_NOT_IN_REFERENTIAL, value=text[:40])
    return None


def _coerce_enum_list(
    raw: str, column: ColumnSpec, workbook_spec: WorkbookSpec, report: Any
) -> list[str] | None:
    """Split a comma-separated cell, dropping empty entries."""
    resolved: list[str] = []
    for piece in raw.split(","):
        text = piece.strip()
        if not text:
            continue
        match = next(
            (a for a in _referential(column, workbook_spec) if text.casefold() == a.casefold()),
            None,
        )
        if match is None:
            report(IssueCode.VALUE_NOT_IN_REFERENTIAL, value=text[:40])
            return None
        resolved.append(match)
    return resolved
