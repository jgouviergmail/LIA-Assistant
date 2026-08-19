"""Unit tests for the workbook writer.

Several assertions read the **emitted XML** rather than openpyxl's Python API.
That is deliberate: two of the attributes this writer depends on have inverted
meaning, and a reader of the Python code would reasonably "fix" them the wrong
way round. Measured on 2026-08-18 against real Excel:

- ``sheetProtection`` booleans mean *blocked* when ``1``. Leaving them at
  openpyxl's defaults protects a sheet **and forbids inserting a row** — that
  is, forbids adding a model, the whole point of the import.
- ``showDropDown="1"`` **hides** the in-cell arrow. Setting the attribute that
  looks like "show the dropdown" removes every dropdown from the workbook.
"""

from __future__ import annotations

import io
import zipfile
from decimal import Decimal

import openpyxl
import pytest

from src.infrastructure.tabular_io.spec import ColumnSpec, SheetSpec, WorkbookSpec
from src.infrastructure.tabular_io.writer import build_workbook

SHEET = SheetSpec(
    name="Models",
    title_key="title.models",
    key_column="model_name",
    columns=(
        ColumnSpec(key="model_name", label_key="l.name", kind="text", block="identity", width=30),
        ColumnSpec(key="provider", label_key="l.provider", kind="enum", referential="PROVIDER"),
        ColumnSpec(key="active", label_key="l.active", kind="boolean"),
        ColumnSpec(key="max_tokens", label_key="l.max", kind="integer"),
        ColumnSpec(
            key="price", label_key="l.price", kind="decimal", decimals=6, minimum=Decimal("0")
        ),
        ColumnSpec(key="status", label_key="l.status", kind="text", editable=False),
    ),
)

SPEC = WorkbookSpec(
    sheets=(SHEET,),
    referentials={"PROVIDER": ("openai", "anthropic")},
    schema_version=3,
)

ROWS = [
    {
        "model_name": "gpt-4.1-mini",
        "provider": "openai",
        "active": True,
        "max_tokens": 1000,
        "price": Decimal("0.400000"),
        "status": "ok",
    },
    {
        "model_name": "claude-x",
        "provider": "anthropic",
        "active": False,
        "max_tokens": 2000,
        "price": None,
        "status": "",
    },
]

LABELS = {
    "l.name": "Nom",
    "l.provider": "Fournisseur",
    "l.active": "Actif",
    "l.max": "Jetons max",
    "l.price": "Prix",
    "l.status": "Statut",
    "title.models": "Modèles",
    "sheet.notice": "Notice",
    "sheet.referentials": "Référentiels",
    "sheet.metadata": "Métadonnées",
    "boolean.true": "VRAI",
    "boolean.false": "FAUX",
}


def _build() -> bytes:
    return build_workbook(
        SPEC,
        {"Models": ROWS},
        notice=["Première ligne de notice", "Deuxième ligne"],
        labels=LABELS,
        metadata={"exported_by": "admin@lia"},
    )


@pytest.fixture(scope="module")
def blob() -> bytes:
    return _build()


@pytest.fixture(scope="module")
def workbook(blob: bytes) -> openpyxl.Workbook:
    return openpyxl.load_workbook(io.BytesIO(blob))


def _sheet_xml(blob: bytes, index: int) -> str:
    with zipfile.ZipFile(io.BytesIO(blob)) as archive:
        return archive.read(f"xl/worksheets/sheet{index}.xml").decode()


def _models_xml(blob: bytes) -> str:
    """Return the XML of the sheet carrying the data validations."""
    for index in range(1, 6):
        try:
            xml = _sheet_xml(blob, index)
        except KeyError:
            continue
        if "<dataValidation " in xml:
            return xml
    raise AssertionError("no sheet carries a data validation")


@pytest.mark.unit
class TestInvertedOpenpyxlFlags:
    """Both are inverted; both are load-bearing; neither is obvious in code."""

    def test_protection_leaves_row_insertion_allowed(self, blob: bytes) -> None:
        xml = _models_xml(blob)
        assert (
            'insertRows="0"' in xml
        ), "insertRows=1 means BLOCKED: an administrator could not add a model"

    def test_protection_leaves_filtering_and_sorting_allowed(self, blob: bytes) -> None:
        xml = _models_xml(blob)
        assert 'autoFilter="0"' in xml and 'sort="0"' in xml

    def test_protection_leaves_row_deletion_and_formatting_allowed(self, blob: bytes) -> None:
        xml = _models_xml(blob)
        for attribute in ("deleteRows", "formatCells", "formatColumns", "formatRows"):
            assert f'{attribute}="0"' in xml

    def test_the_sheet_is_still_protected(self, blob: bytes) -> None:
        assert 'sheet="1"' in _models_xml(blob)

    def test_dropdown_arrows_are_visible(self, blob: bytes) -> None:
        xml = _models_xml(blob)
        assert (
            'showDropDown="1"' not in xml
        ), "showDropDown=1 HIDES the arrow — every dropdown would disappear"

    def test_no_protection_password_is_set(self, blob: bytes) -> None:
        """A password is bypassed in seconds and only creates false confidence."""
        assert "password" not in _models_xml(blob)


@pytest.mark.unit
class TestLayout:
    def test_every_sheet_is_present(self, workbook: openpyxl.Workbook) -> None:
        assert workbook.sheetnames == ["Notice", "Models", "Référentiels", "Métadonnées"]

    def test_row_one_carries_the_technical_keys(self, workbook: openpyxl.Workbook) -> None:
        sheet = workbook["Models"]
        assert [sheet.cell(row=1, column=i).value for i in range(1, 7)] == list(SHEET.keys)

    def test_row_one_is_hidden(self, workbook: openpyxl.Workbook) -> None:
        assert workbook["Models"].row_dimensions[1].hidden is True

    def test_row_two_carries_the_translated_labels(self, workbook: openpyxl.Workbook) -> None:
        sheet = workbook["Models"]
        assert sheet.cell(row=2, column=1).value == "Nom"
        assert sheet.cell(row=2, column=2).value == "Fournisseur"

    def test_data_starts_on_row_three(self, workbook: openpyxl.Workbook) -> None:
        assert workbook["Models"].cell(row=3, column=1).value == "gpt-4.1-mini"

    def test_the_key_column_and_both_headers_stay_visible(
        self, workbook: openpyxl.Workbook
    ) -> None:
        assert workbook["Models"].freeze_panes == "B3"

    def test_the_referential_sheet_is_hidden(self, workbook: openpyxl.Workbook) -> None:
        assert workbook["Référentiels"].sheet_state == "hidden"

    def test_the_notice_carries_the_supplied_lines(self, workbook: openpyxl.Workbook) -> None:
        notice = workbook["Notice"]
        assert notice.cell(row=1, column=1).value == "Première ligne de notice"

    def test_metadata_records_the_schema_version(self, workbook: openpyxl.Workbook) -> None:
        sheet = workbook["Métadonnées"]
        pairs = {
            sheet.cell(row=r, column=1).value: sheet.cell(row=r, column=2).value
            for r in range(1, sheet.max_row + 1)
        }
        assert pairs["sheet_schema_version"] == "3"
        assert pairs["exported_by"] == "admin@lia"


@pytest.mark.unit
class TestCellContract:
    def test_editable_cells_are_unlocked(self, workbook: openpyxl.Workbook) -> None:
        assert workbook["Models"].cell(row=3, column=1).protection.locked is False

    def test_read_only_cells_stay_locked(self, workbook: openpyxl.Workbook) -> None:
        status_column = SHEET.keys.index("status") + 1
        assert workbook["Models"].cell(row=3, column=status_column).protection.locked is True

    def test_header_cells_stay_locked(self, workbook: openpyxl.Workbook) -> None:
        assert workbook["Models"].cell(row=2, column=1).protection.locked is True

    def test_booleans_use_the_localized_words(self, workbook: openpyxl.Workbook) -> None:
        sheet = workbook["Models"]
        assert sheet.cell(row=3, column=3).value == "VRAI"
        assert sheet.cell(row=4, column=3).value == "FAUX"

    def test_decimals_are_written_as_numbers_with_their_scale(
        self, workbook: openpyxl.Workbook
    ) -> None:
        cell = workbook["Models"].cell(row=3, column=5)
        assert cell.value == 0.4
        assert cell.number_format == "0.000000"

    def test_an_absent_decimal_leaves_the_cell_empty(self, workbook: openpyxl.Workbook) -> None:
        assert workbook["Models"].cell(row=4, column=5).value is None

    def test_an_empty_string_is_written_as_an_empty_cell(self, workbook: openpyxl.Workbook) -> None:
        """Excel returns an empty cell as None; writing "" would make the round
        trip report a phantom change on every such row."""
        status_column = SHEET.keys.index("status") + 1
        assert workbook["Models"].cell(row=4, column=status_column).value is None

    def test_a_formula_shaped_value_is_neutralized(self) -> None:
        """A stored value starting with = must not become a live formula."""
        blob = build_workbook(
            SPEC,
            {"Models": [{**ROWS[0], "model_name": "=cmd|' /c calc'!A0"}]},
            notice=[],
            labels=LABELS,
            metadata={},
        )
        value = openpyxl.load_workbook(io.BytesIO(blob))["Models"].cell(row=3, column=1).value
        assert isinstance(value, str) and not value.startswith("=")

    def test_a_negative_number_is_not_mistaken_for_a_formula(self) -> None:
        blob = build_workbook(
            SPEC,
            {"Models": [{**ROWS[0], "status": "-5.2"}]},
            notice=[],
            labels=LABELS,
            metadata={},
        )
        status_column = SHEET.keys.index("status") + 1
        value = (
            openpyxl.load_workbook(io.BytesIO(blob))["Models"]
            .cell(row=3, column=status_column)
            .value
        )
        assert value == "-5.2"


@pytest.mark.unit
class TestValidations:
    def test_enum_columns_get_a_dropdown_on_their_referential(
        self, workbook: openpyxl.Workbook
    ) -> None:
        validations = workbook["Models"].data_validations.dataValidation
        formulas = {v.formula1 for v in validations if v.type == "list"}
        assert "=LST_PROVIDER" in formulas

    def test_boolean_columns_get_a_dropdown(self, workbook: openpyxl.Workbook) -> None:
        validations = workbook["Models"].data_validations.dataValidation
        formulas = {v.formula1 for v in validations if v.type == "list"}
        assert any("BOOL" in str(f) for f in formulas)

    def test_decimal_columns_carry_their_lower_bound(self, workbook: openpyxl.Workbook) -> None:
        validations = workbook["Models"].data_validations.dataValidation
        numeric = [v for v in validations if v.type == "decimal"]
        assert numeric and numeric[0].formula1 == "0"

    def test_read_only_columns_get_no_dropdown(self, workbook: openpyxl.Workbook) -> None:
        """A dropdown on a column the reader ignores would promise an effect."""
        status_letter = openpyxl.utils.get_column_letter(SHEET.keys.index("status") + 1)
        for validation in workbook["Models"].data_validations.dataValidation:
            assert status_letter not in str(validation.sqref)

    def test_every_referential_is_written_as_a_defined_name(
        self, workbook: openpyxl.Workbook
    ) -> None:
        assert "LST_PROVIDER" in workbook.defined_names


@pytest.mark.unit
class TestEmptyAndEdgeCases:
    def test_a_sheet_with_no_row_still_carries_its_headers(self) -> None:
        blob = build_workbook(SPEC, {"Models": []}, notice=[], labels=LABELS, metadata={})
        sheet = openpyxl.load_workbook(io.BytesIO(blob))["Models"]
        assert sheet.cell(row=1, column=1).value == "model_name"

    def test_a_missing_value_is_written_as_an_empty_cell(self) -> None:
        blob = build_workbook(
            SPEC, {"Models": [{"model_name": "only-key"}]}, notice=[], labels=LABELS, metadata={}
        )
        sheet = openpyxl.load_workbook(io.BytesIO(blob))["Models"]
        assert sheet.cell(row=3, column=1).value == "only-key"
        assert sheet.cell(row=3, column=2).value is None

    def test_an_unknown_label_key_falls_back_to_the_column_key(self) -> None:
        """A missing translation must not produce an empty header."""
        blob = build_workbook(SPEC, {"Models": ROWS}, notice=[], labels={}, metadata={})
        sheet = openpyxl.load_workbook(io.BytesIO(blob))["Models"]
        assert sheet.cell(row=2, column=1).value == "model_name"

    def test_a_sheet_absent_from_the_data_is_written_empty(self) -> None:
        blob = build_workbook(SPEC, {}, notice=[], labels=LABELS, metadata={})
        sheet = openpyxl.load_workbook(io.BytesIO(blob))["Models"]
        assert sheet.max_row >= 2
