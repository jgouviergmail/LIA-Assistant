"""Unit tests for the workbook reader.

Every behaviour pinned here traces to something measured on 2026-08-18, not to
something imagined:

- a formula typed in Excel comes back as the string ``'=0.1+0.2'`` — so
  formulas are rejectable deterministically, and ``data_only=True`` is refused
  (it yields ``None`` on a workbook Excel never opened);
- a price typed as text comes back as ``'0,7'`` — coercion is mandatory;
- ``max_row`` reads 500 for four rows of data — no count may come from it;
- an empty cell reads back as ``None`` while the writer was handed ``""`` —
  both must mean the same thing, or 122 rows report a phantom change.
"""

from __future__ import annotations

import io
import zipfile
from decimal import Decimal

import openpyxl
import pytest

from src.infrastructure.tabular_io.reader import parse_workbook
from src.infrastructure.tabular_io.report import IssueCode
from src.infrastructure.tabular_io.spec import ColumnSpec, SheetSpec, WorkbookSpec
from src.infrastructure.tabular_io.writer import DATA_TOP_ROW

SHEET = SheetSpec(
    name="Models",
    title_key="t",
    key_column="model_name",
    columns=(
        ColumnSpec(key="model_name", label_key="l", kind="text", required=True),
        ColumnSpec(key="provider", label_key="l", kind="enum", referential="PROVIDER"),
        ColumnSpec(key="active", label_key="l", kind="boolean"),
        ColumnSpec(key="max_tokens", label_key="l", kind="integer", minimum=Decimal("0")),
        ColumnSpec(key="price", label_key="l", kind="decimal", decimals=6, minimum=Decimal("0")),
        ColumnSpec(key="window", label_key="l", kind="time_hhmm"),
        ColumnSpec(key="tags", label_key="l", kind="enum_list", referential="TAG"),
        ColumnSpec(key="status", label_key="l", kind="text", editable=False),
    ),
)
SPEC = WorkbookSpec(
    sheets=(SHEET,),
    referentials={"PROVIDER": ("openai", "anthropic"), "TAG": ("a", "b")},
    schema_version=1,
)

LIMITS = {"max_rows": 500, "max_files": 100, "max_decompressed_bytes": 10_000_000}


def _workbook(rows: list[dict[str, object]], *, version: str | None = "1") -> bytes:
    """Build a workbook by hand so tests control the exact cell contents."""
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    sheet = workbook.create_sheet("Models")
    for index, key in enumerate(SHEET.keys, start=1):
        sheet.cell(row=1, column=index, value=key)
        sheet.cell(row=2, column=index, value=key.upper())
    for offset, row in enumerate(rows):
        for index, key in enumerate(SHEET.keys, start=1):
            if key in row:
                sheet.cell(row=DATA_TOP_ROW + offset, column=index, value=row[key])
    if version is not None:
        meta = workbook.create_sheet("Metadata")
        meta.cell(row=1, column=1, value="sheet_schema_version")
        meta.cell(row=1, column=2, value=version)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _parse(rows: list[dict[str, object]], **overrides: object):
    return parse_workbook(SPEC, _workbook(rows), **{**LIMITS, **overrides})  # type: ignore[arg-type]


def _codes(result) -> set[IssueCode]:
    return {issue.code for issue in result.issues}


@pytest.mark.unit
class TestNominalParsing:
    def test_a_well_formed_row_is_parsed_into_typed_values(self) -> None:
        result = _parse(
            [
                {
                    "model_name": "gpt",
                    "provider": "openai",
                    "active": "VRAI",
                    "max_tokens": 1000,
                    "price": 0.4,
                    "window": "01:00",
                    "tags": "a,b",
                }
            ]
        )
        assert not result.issues
        row = result.rows("Models")[0]
        assert row.values["model_name"] == "gpt"
        assert row.values["active"] is True
        assert row.values["max_tokens"] == 1000
        assert row.values["price"] == Decimal("0.4")
        assert row.values["tags"] == ["a", "b"]

    def test_each_row_reports_the_worksheet_row_it_came_from(self) -> None:
        result = _parse([{"model_name": "a"}, {"model_name": "b"}])
        assert [row.row_number for row in result.rows("Models")] == [3, 4]

    def test_read_only_columns_are_ignored(self) -> None:
        result = _parse([{"model_name": "gpt", "status": "whatever the admin typed"}])
        assert "status" not in result.rows("Models")[0].values


@pytest.mark.unit
class TestFormulasAreRefused:
    def test_a_formula_in_a_text_cell_is_reported(self) -> None:
        result = _parse([{"model_name": "=1+2"}])
        assert IssueCode.FORMULA_REJECTED in _codes(result)

    def test_a_formula_in_a_numeric_cell_is_reported(self) -> None:
        result = _parse([{"model_name": "gpt", "price": "=SUM(A1:A2)"}])
        assert IssueCode.FORMULA_REJECTED in _codes(result)

    def test_the_offending_cell_is_named(self) -> None:
        result = _parse([{"model_name": "=1+2"}])
        issue = next(i for i in result.issues if i.code is IssueCode.FORMULA_REJECTED)
        assert issue.cell == "A3" and issue.sheet == "Models"


@pytest.mark.unit
class TestEmptinessAndCounting:
    def test_fully_empty_rows_are_skipped(self) -> None:
        result = _parse([{"model_name": "a"}, {}, {"model_name": "b"}])
        assert [row.values["model_name"] for row in result.rows("Models")] == ["a", "b"]

    def test_a_row_of_blank_strings_counts_as_empty(self) -> None:
        result = _parse([{"model_name": "   ", "provider": ""}])
        assert result.rows("Models") == ()

    def test_no_count_is_derived_from_max_row(self) -> None:
        """The writer pre-formats spare rows, inflating max_row to hundreds."""
        blob = _workbook([{"model_name": "only-one"}])
        book = openpyxl.load_workbook(io.BytesIO(blob))
        book["Models"].cell(row=400, column=1).number_format = "0.00"
        buffer = io.BytesIO()
        book.save(buffer)

        result = parse_workbook(SPEC, buffer.getvalue(), **LIMITS)  # type: ignore[arg-type]

        assert len(result.rows("Models")) == 1

    def test_an_empty_string_and_an_absent_cell_are_the_same_value(self) -> None:
        with_empty = _parse([{"model_name": "gpt", "status": ""}])
        without = _parse([{"model_name": "gpt"}])
        assert with_empty.rows("Models")[0].values == without.rows("Models")[0].values


@pytest.mark.unit
class TestCoercion:
    @pytest.mark.parametrize(
        ("raw", "expected"),
        [
            (0.4, Decimal("0.4")),
            ("0,7", Decimal("0.7")),
            ("1 234,50", Decimal("1234.50")),
            ("  2.5  ", Decimal("2.5")),
            (1, Decimal("1")),
        ],
    )
    def test_numbers_survive_the_shapes_excel_really_returns(
        self, raw: object, expected: Decimal
    ) -> None:
        result = _parse([{"model_name": "gpt", "price": raw}])
        assert result.rows("Models")[0].values["price"] == expected

    @pytest.mark.parametrize("raw", ["VRAI", "true", " Oui ", "1", True])
    def test_truthy_words_are_accepted(self, raw: object) -> None:
        result = _parse([{"model_name": "gpt", "active": raw}])
        assert result.rows("Models")[0].values["active"] is True

    @pytest.mark.parametrize("raw", ["FAUX", "false", "NON", "0", False])
    def test_falsy_words_are_accepted(self, raw: object) -> None:
        result = _parse([{"model_name": "gpt", "active": raw}])
        assert result.rows("Models")[0].values["active"] is False

    def test_an_unreadable_boolean_is_reported(self) -> None:
        result = _parse([{"model_name": "gpt", "active": "peut-être"}])
        assert IssueCode.NOT_A_BOOLEAN in _codes(result)

    def test_an_unreadable_number_is_reported(self) -> None:
        result = _parse([{"model_name": "gpt", "price": "gratuit"}])
        assert IssueCode.NOT_A_NUMBER in _codes(result)

    def test_surrounding_spaces_are_trimmed_from_text(self) -> None:
        result = _parse([{"model_name": "  gpt-4.1  "}])
        assert result.rows("Models")[0].values["model_name"] == "gpt-4.1"

    def test_enum_matching_ignores_case_and_spacing(self) -> None:
        result = _parse([{"model_name": "gpt", "provider": " OpenAI "}])
        assert result.rows("Models")[0].values["provider"] == "openai"

    def test_an_enum_list_tolerates_spacing_and_empty_entries(self) -> None:
        result = _parse([{"model_name": "gpt", "tags": " a , , b "}])
        assert result.rows("Models")[0].values["tags"] == ["a", "b"]


@pytest.mark.unit
class TestConstraints:
    def test_more_decimals_than_declared_is_refused_never_rounded(self) -> None:
        result = _parse([{"model_name": "gpt", "price": "0.1234567"}])
        assert IssueCode.TOO_MANY_DECIMALS in _codes(result)

    def test_exactly_the_declared_scale_is_accepted(self) -> None:
        result = _parse([{"model_name": "gpt", "price": "0.123456"}])
        assert not result.issues

    def test_a_value_below_the_minimum_is_refused(self) -> None:
        result = _parse([{"model_name": "gpt", "price": "-1"}])
        assert IssueCode.OUT_OF_RANGE in _codes(result)

    def test_a_value_outside_the_referential_is_refused(self) -> None:
        result = _parse([{"model_name": "gpt", "provider": "acme"}])
        assert IssueCode.VALUE_NOT_IN_REFERENTIAL in _codes(result)

    def test_a_malformed_clock_time_is_refused(self) -> None:
        result = _parse([{"model_name": "gpt", "window": "25:00"}])
        assert IssueCode.NOT_A_TIME in _codes(result)

    @pytest.mark.parametrize("raw", ["01:00", "23:59", "00:00"])
    def test_valid_clock_times_are_accepted(self, raw: str) -> None:
        result = _parse([{"model_name": "gpt", "window": raw}])
        assert result.rows("Models")[0].values["window"] == raw

    def test_a_missing_required_value_is_reported(self) -> None:
        result = _parse([{"model_name": "gpt"}, {"provider": "openai"}])
        assert IssueCode.KEY_MISSING in _codes(result)

    def test_a_duplicate_key_is_reported(self) -> None:
        result = _parse([{"model_name": "twice"}, {"model_name": "twice"}])
        assert IssueCode.DUPLICATE_KEY in _codes(result)

    def test_more_rows_than_allowed_is_refused(self) -> None:
        result = _parse([{"model_name": f"m{i}"} for i in range(5)], max_rows=3)
        assert IssueCode.TOO_MANY_ROWS in _codes(result)


@pytest.mark.unit
class TestStructure:
    def test_columns_are_resolved_by_key_not_by_position(self) -> None:
        """Reordering columns in Excel must be a non-event."""
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        sheet = workbook.create_sheet("Models")
        reordered = ["price", "model_name", "provider"]
        for index, key in enumerate(reordered, start=1):
            sheet.cell(row=1, column=index, value=key)
        sheet.cell(row=DATA_TOP_ROW, column=1, value=0.5)
        sheet.cell(row=DATA_TOP_ROW, column=2, value="gpt")
        sheet.cell(row=DATA_TOP_ROW, column=3, value="openai")
        buffer = io.BytesIO()
        workbook.save(buffer)

        result = parse_workbook(SPEC, buffer.getvalue(), **LIMITS)  # type: ignore[arg-type]

        row = result.rows("Models")[0]
        assert row.values["model_name"] == "gpt" and row.values["price"] == Decimal("0.5")

    def test_an_unknown_extra_column_is_ignored(self) -> None:
        workbook = openpyxl.load_workbook(io.BytesIO(_workbook([{"model_name": "gpt"}])))
        sheet = workbook["Models"]
        sheet.cell(row=1, column=20, value="notes_de_l_admin")
        sheet.cell(row=DATA_TOP_ROW, column=20, value="rappel perso")
        buffer = io.BytesIO()
        workbook.save(buffer)

        result = parse_workbook(SPEC, buffer.getvalue(), **LIMITS)  # type: ignore[arg-type]

        assert not result.issues

    def test_a_missing_required_column_is_reported(self) -> None:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        sheet = workbook.create_sheet("Models")
        sheet.cell(row=1, column=1, value="provider")
        buffer = io.BytesIO()
        workbook.save(buffer)

        result = parse_workbook(SPEC, buffer.getvalue(), **LIMITS)  # type: ignore[arg-type]

        assert IssueCode.COLUMN_MISSING in _codes(result)

    def test_a_missing_sheet_is_reported(self) -> None:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        workbook.create_sheet("SomethingElse")
        buffer = io.BytesIO()
        workbook.save(buffer)

        result = parse_workbook(SPEC, buffer.getvalue(), **LIMITS)  # type: ignore[arg-type]

        assert IssueCode.SHEET_MISSING in _codes(result)

    def test_a_schema_version_mismatch_is_refused(self) -> None:
        result = parse_workbook(SPEC, _workbook([{"model_name": "a"}], version="99"), **LIMITS)  # type: ignore[arg-type]
        assert IssueCode.SCHEMA_VERSION_MISMATCH in _codes(result)

    def test_an_absent_metadata_sheet_is_tolerated(self) -> None:
        """A workbook rebuilt by hand keeps working; only a WRONG version fails."""
        result = parse_workbook(SPEC, _workbook([{"model_name": "a"}], version=None), **LIMITS)  # type: ignore[arg-type]
        assert IssueCode.SCHEMA_VERSION_MISMATCH not in _codes(result)


@pytest.mark.unit
class TestHostileInput:
    @pytest.mark.parametrize("payload", [b"", b"not a zip at all", b"PK\x03\x04garbage"])
    def test_anything_that_is_not_a_workbook_is_refused(self, payload: bytes) -> None:
        result = parse_workbook(SPEC, payload, **LIMITS)  # type: ignore[arg-type]
        assert IssueCode.NOT_A_WORKBOOK in _codes(result)
        assert result.rows("Models") == ()

    def test_an_archive_over_its_decompressed_budget_is_refused(self) -> None:
        blob = _workbook([{"model_name": f"m{i}"} for i in range(50)])
        result = parse_workbook(SPEC, blob, max_rows=500, max_files=100, max_decompressed_bytes=10)
        assert IssueCode.ARCHIVE_TOO_LARGE in _codes(result)

    def test_an_archive_with_too_many_members_is_refused(self) -> None:
        blob = _workbook([{"model_name": "a"}])
        result = parse_workbook(SPEC, blob, max_rows=500, max_files=1, max_decompressed_bytes=10**7)
        assert IssueCode.ARCHIVE_TOO_LARGE in _codes(result)

    def test_a_refused_archive_yields_no_row_at_all(self) -> None:
        result = parse_workbook(SPEC, b"nope", **LIMITS)  # type: ignore[arg-type]
        assert result.rows("Models") == () and result.has_blocking_issues

    def test_the_workbook_is_never_opened_with_cached_values(self) -> None:
        """``data_only=True`` returns None on a file Excel never opened, making
        the result depend on which tool last saved it."""
        blob = _workbook([{"model_name": "gpt", "price": "=1+1"}])
        with zipfile.ZipFile(io.BytesIO(blob)) as archive:
            assert "xl/calcChain.xml" not in archive.namelist()
        result = parse_workbook(SPEC, blob, **LIMITS)  # type: ignore[arg-type]
        assert IssueCode.FORMULA_REJECTED in _codes(result)


@pytest.mark.unit
class TestNonFiniteNumbers:
    """``Decimal`` accepts NaN and Infinity; a tariff must not.

    Both parse cleanly and both slip past a ``< minimum`` check (comparisons
    with NaN are always false), so without an explicit guard they reached the
    scale computation and crashed the whole import with a TypeError — a 500
    where the administrator deserved "cell C7 is not a number".
    """

    @pytest.mark.parametrize("raw", ["NaN", "nan", "inf", "-inf", "Infinity", "sNaN"])
    def test_non_finite_values_are_reported_not_crashed(self, raw: str) -> None:
        result = _parse([{"model_name": "gpt", "price": raw}])
        assert IssueCode.NOT_A_NUMBER in _codes(result)

    @pytest.mark.parametrize("raw", ["NaN", "inf"])
    def test_non_finite_values_never_reach_the_parsed_row(self, raw: str) -> None:
        result = _parse([{"model_name": "gpt", "price": raw}])
        assert result.rows("Models")[0].values["price"] is None

    def test_a_non_finite_integer_is_also_refused(self) -> None:
        result = _parse([{"model_name": "gpt", "max_tokens": "inf"}])
        assert IssueCode.NOT_A_NUMBER in _codes(result)


@pytest.mark.unit
class TestGroupingKeys:
    """A detail sheet repeats its parent key on purpose."""

    def test_repeated_keys_are_refused_on_an_identifying_sheet(self) -> None:
        result = _parse([{"model_name": "same"}, {"model_name": "same"}])
        assert IssueCode.DUPLICATE_KEY in _codes(result)

    def test_repeated_keys_are_accepted_on_a_grouping_sheet(self) -> None:
        grouping = SheetSpec(
            name="Models",
            title_key="t",
            key_column="model_name",
            key_is_unique=False,
            columns=SHEET.columns,
        )
        spec = WorkbookSpec(
            sheets=(grouping,),
            referentials={"PROVIDER": ("openai",), "TAG": ("a",)},
            schema_version=1,
        )

        result = parse_workbook(spec, _workbook([{"model_name": "same"}, {"model_name": "same"}]), **LIMITS)  # type: ignore[arg-type]

        assert IssueCode.DUPLICATE_KEY not in _codes(result)
        assert len(result.rows("Models")) == 2


@pytest.mark.unit
class TestDerivedValues:
    """Read-only columns are not domain input, but they are still readable.

    The import needs the fingerprint the export wrote; the diff needs to ignore
    it as a value. Keeping derived cells apart from editable ones gives both.
    """

    def test_read_only_columns_land_in_derived_not_in_values(self) -> None:
        result = _parse([{"model_name": "gpt", "status": "computed"}])
        row = result.rows("Models")[0]
        assert "status" not in row.values
        assert row.derived["status"] == "computed"

    def test_a_derived_cell_is_coerced_like_any_other(self) -> None:
        result = _parse([{"model_name": "gpt", "status": "  spaced  "}])
        assert result.rows("Models")[0].derived["status"] == "spaced"

    def test_an_absent_derived_cell_reads_as_none(self) -> None:
        result = _parse([{"model_name": "gpt"}])
        assert result.rows("Models")[0].derived["status"] is None

    def test_a_derived_cell_never_produces_an_issue(self) -> None:
        """The admin cannot fix what they cannot edit: complaining is pointless."""
        result = _parse([{"model_name": "gpt", "status": "=1+2"}])
        assert not result.issues
